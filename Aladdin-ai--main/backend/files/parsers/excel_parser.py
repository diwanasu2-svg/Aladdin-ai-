"""Excel parser — reads .xlsx/.xls sheets."""
from __future__ import annotations
import asyncio, io, logging
from typing import Any, Dict
log = logging.getLogger(__name__)

try:
    import openpyxl
    _XL = True
except ImportError:
    _XL = False
    log.warning("openpyxl not installed")


class ExcelParser:
    @property
    def available(self): return _XL

    async def parse(self, data: bytes, max_rows: int = 1000) -> Dict[str, Any]:
        if not _XL:
            return {"error": "openpyxl not installed", "sheets": []}
        def _run():
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows: break
                    rows.append([str(c) if c is not None else "" for c in row])
                sheets.append({"name": name, "rows": rows,
                                "row_count": ws.max_row, "col_count": ws.max_column})
            return {"sheets": sheets, "sheet_count": len(sheets)}
        return await asyncio.get_running_loop().run_in_executor(None, _run)
